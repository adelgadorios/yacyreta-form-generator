from tkinter import *
from tkinter import ttk
from openpyxl import *
import openpyxl
import calendar
import tkinter as tk
import time
import shutil
import tkinter.messagebox
from openpyxl.styles import Font
from openpyxl.styles import colors
import os
import datetime
from openpyxl.drawing.image import Image
root = Tk()
n = StringVar()
lg = StringVar()
dp = StringVar()


def error():
    lg.set("")
    n.set("")
    dp.set("")

h=open("backup.txt","r+")
h.close
if os.stat("backup.txt").st_size > 0:
    try:
        with open('backup.txt', 'r') as f:
            x = f.readlines()
            lg.set(x[0])
            n.set(x[1])
            dp.set(x[2])

    except:
        error()

root.geometry("640x450")
root.resizable(width=False, height=False)
root.iconbitmap( default="assets/minilogo.ico")
photo = PhotoImage(file="assets/descarga.png")
Labelo = Label(root,image=photo).place(x=220, y=-50, relwidth=1, relheight=1)
root.title('Formulario')
root.state("normal")
heading = Label(root,text="Bienvenido al Programa de Formularios" ,fg="black"  ,font=("arial",20 )).pack()
motivos= {
"REPOSO":1,
"VACACIONES":2,
"COMISION DE SERVICIO INTERSEDE":3,
"COMISION DE SERVICIO LOCAL":4,
"ATENCION FAMILIAR":5,
"EXAMEN":6,
"MATRIMONIO":7,
"MATERNIDAD":8,
"PATERNIDAD":9,
"DUELO (PADRES, CONYUGE, HIJOS)":10,
"DUELO (OTROS)":11,
"LICENCIA SINDICAL":12,
"PARTICULAR":13,
"LACTANCIA":14,
"CAPACITACION":15,
"LLEGADA TARDIA":16,
"SIN REGISTRO DE ENTRADA":17,
"SIN REGISTRO DE SALIDA":18,
"SALIDA ANTES DE HORA":19,
"COMPENSACION":20,
"OTROS (INDICAR OBSERVACION)":21

}


Label1 = ttk.Label(root, text="Ingrese el numero de legajo:").place(x=0, y=50)

entry_box1 = ttk.Entry(root, width=40, textvariable=lg).place(x=200, y=50)

Label2 = ttk.Label(root, text="Ingrese Nombre y apellido:").place(x=0, y=90)

entry_box2 = ttk.Entry(root, width=40, textvariable=n).place(x=200, y=90)

Label3 = ttk.Label(root, text="Ingrese Dependencia:").place(x=0, y=130)

entry_box3 = ttk.Entry(root, width=40, textvariable=dp).place(x=200, y=130)


Label7 = ttk.Label(root, text="Ingrese  observacion:").place(x=0, y=170)
ci = StringVar()
entry_box7 = ttk.Entry(root, width=40, textvariable=ci).place(x=200, y=170)

Label6 = Label(root, text="Ingrese  hora inicial:").place(x=0, y=210)
h1 = StringVar()
entry_box6 = ttk.Entry(root, width=40, textvariable=h1).place(x=200, y=210)

Label23 = Label(root, text="Ingrese hora final:").place(x=0, y=250)
h2 = StringVar()
entry_box23 = ttk.Entry(root, width=40, textvariable=h2).place(x=200, y=250)

Label4 = ttk.Label(root, text="Elija codigo del motivo:").place(x=0, y=290)
cm = StringVar(root)
cm.set(motivos.values())
motivo = ttk.OptionMenu(root,  cm, *motivos.keys()).place(x=200, y=290)



#codigo de fecha#

class MyDatePicker(tk.Toplevel):
    """
    Description:
        A tkinter GUI date picker.
    """

    def __init__(self, parent=None):
        """
        Description:
            When instantiating in parent module/widget/Gui, pass in 'self' as argument.
            Ex:
                a = MyDatePicker(self)

        :param parent: parent instance.
        """

        super().__init__()
        self.parent = parent
        self.title("Fechas")
        self.resizable(0, 0)
        self.geometry("+250+10")
        self.init_frames()
        self.init_needed_vars()
        self.init_month_year_labels()
        self.init_buttons()
        self.space_between_widgets()
        self.fill_days()
        self.make_calendar()

    def init_frames(self):
        self.frame1 = tk.Frame(self)
        self.frame1.pack()

        self.frame_days = tk.Frame(self)
        self.frame_days.pack()

    def init_needed_vars(self):
        self.month_names = tuple(calendar.month_name)
        self.day_names = tuple(calendar.day_abbr)
        self.year = time.strftime("%Y")
        self.month = time.strftime("%B")

    def init_month_year_labels(self):
        self.year_str_var = tk.StringVar()
        self.month_str_var = tk.StringVar()

        self.year_str_var.set(self.year)
        self.year_lbl = tk.Label(self.frame1, textvariable=self.year_str_var, width=3)
        self.year_lbl.grid(row=0, column=5)

        self.month_str_var.set(self.month)
        self.month_lbl = tk.Label(self.frame1, textvariable=self.month_str_var, width=8)
        self.month_lbl.grid(row=0, column=1)

    def init_buttons(self):
        self.left_yr = ttk.Button(self.frame1, text="←", width=5, command=self.prev_year)
        self.left_yr.grid(row=0, column=4)

        self.right_yr = ttk.Button(self.frame1, text="→", width=5, command=self.next_year)
        self.right_yr.grid(row=0, column=6)

        self.left_mon = ttk.Button(self.frame1, text="←", width=5, command=self.prev_month)
        self.left_mon.grid(row=0, column=0)

        self.right_mon = ttk.Button(self.frame1, text="→", width=5, command=self.next_month)
        self.right_mon.grid(row=0, column=2)

    def space_between_widgets(self):
        self.frame1.grid_columnconfigure(3, minsize=40)

    def prev_year(self):
        self.prev_yr = int(self.year_str_var.get()) - 1
        self.year_str_var.set(self.prev_yr)

        self.make_calendar()

    def next_year(self):
        self.next_yr = int(self.year_str_var.get()) + 1
        self.year_str_var.set(self.next_yr)

        self.make_calendar()

    def prev_month(self):
        index_current_month = int(self.month_names.index(self.month_str_var.get()))
        index_prev_month = index_current_month - 1

        #  index 0 is empty string, use index 12 instead, which is index of December.
        if index_prev_month == 0:
            self.month_str_var.set(self.month_names[12])
        else:
            self.month_str_var.set(self.month_names[index_current_month - 1])

        self.make_calendar()

    def next_month(self):
        index_current_month = int(self.month_names.index(self.month_str_var.get()))

        #  index 13 does not exist, use index 1 instead, which is January.
        try:
            self.month_str_var.set(self.month_names[index_current_month + 1])
        except IndexError:
            self.month_str_var.set(self.month_names[1])

        self.make_calendar()

    def fill_days(self):
        col = 0
        #  Creates days label
        for day in self.day_names:
            self.lbl_day = tk.Label(self.frame_days, text=day)
            self.lbl_day.grid(row=0, column=col)
            col += 1

    def make_calendar(self):
        #  Delete date buttons if already present.
        #  Each button must have its own instance attribute for this to work.
        try:
            for dates in self.m_cal:
                for date in dates:
                    if date == 0:
                        continue

                    self.delete_buttons(date)

        except AttributeError:
            pass

        year = int(self.year_str_var.get())
        month = self.month_names.index(self.month_str_var.get())
        self.m_cal = calendar.monthcalendar(year, month)

        #  build date buttons.
        for dates in self.m_cal:
            row = self.m_cal.index(dates) + 1
            for date in dates:
                col = dates.index(date)

                if date == 0:
                    continue

                self.make_button(str(date), str(row), str(col))

    def make_button(self, date, row, column):
        exec(
            "self.btn_" + date + "= ttk.Button(self.frame_days, text=" + date + ", width=5)\n"
            "self.btn_" + date + ".grid(row=" + row + " , column=" + column + ")\n"
            "self.btn_" + date + ".bind(\"<Button-1>\", self.get_date)"
        )

    def delete_buttons(self, date):
        exec(
            "self.btn_" + str(date) + ".destroy()"
        )

    def get_date(self, clicked=None):
        clicked_button = clicked.widget
        year = self.year_str_var.get()
        month = self.month_names.index(self.month_str_var.get())
        date = clicked_button['text']
        #  Change string format for different date formats.
        global full_date
        full_date = '%02d-%02d-%s' % (date, month, year)
        fecha2l = ttk.Label(root, text=full_date).place(x=350, y=330)
        print(full_date)
        self.destroy()

if __name__ == '__main__':
    def application():
        app = MyDatePicker()

    class MyDatePicker2(tk.Toplevel):
        """
        Description:
            A tkinter GUI date picker.
        """

        def __init__(self, parent=None):
            """
            Description:
                When instantiating in parent module/widget/Gui, pass in 'self' as argument.
                Ex:
                    a = MyDatePicker(self)

            :param parent: parent instance.
            """

            super().__init__()
            self.parent = parent
            self.title("Date Picker")
            self.resizable(0, 0)
            self.geometry("+250+10")
            self.init_frames()
            self.init_needed_vars()
            self.init_month_year_labels()
            self.init_buttons()
            self.space_between_widgets()
            self.fill_days()
            self.make_calendar()

        def init_frames(self):
            self.frame1 = tk.Frame(self)
            self.frame1.pack()

            self.frame_days = tk.Frame(self)
            self.frame_days.pack()

        def init_needed_vars(self):
            self.month_names = tuple(calendar.month_name)
            self.day_names = tuple(calendar.day_abbr)
            self.year = time.strftime("%Y")
            self.month = time.strftime("%B")

        def init_month_year_labels(self):
            self.year_str_var = tk.StringVar()
            self.month_str_var = tk.StringVar()

            self.year_str_var.set(self.year)
            self.year_lbl = tk.Label(self.frame1, textvariable=self.year_str_var, width=3)
            self.year_lbl.grid(row=0, column=5)

            self.month_str_var.set(self.month)
            self.month_lbl = tk.Label(self.frame1, textvariable=self.month_str_var, width=8)
            self.month_lbl.grid(row=0, column=1)

        def init_buttons(self):
            self.left_yr = ttk.Button(self.frame1, text="←", width=5, command=self.prev_year)
            self.left_yr.grid(row=0, column=4)

            self.right_yr = ttk.Button(self.frame1, text="→", width=5, command=self.next_year)
            self.right_yr.grid(row=0, column=6)

            self.left_mon = ttk.Button(self.frame1, text="←", width=5, command=self.prev_month)
            self.left_mon.grid(row=0, column=0)

            self.right_mon = ttk.Button(self.frame1, text="→", width=5, command=self.next_month)
            self.right_mon.grid(row=0, column=2)

        def space_between_widgets(self):
            self.frame1.grid_columnconfigure(3, minsize=40)

        def prev_year(self):
            self.prev_yr = int(self.year_str_var.get()) - 1
            self.year_str_var.set(self.prev_yr)

            self.make_calendar()

        def next_year(self):
            self.next_yr = int(self.year_str_var.get()) + 1
            self.year_str_var.set(self.next_yr)

            self.make_calendar()

        def prev_month(self):
            index_current_month = int(self.month_names.index(self.month_str_var.get()))
            index_prev_month = index_current_month - 1

            #  index 0 is empty string, use index 12 instead, which is index of December.
            if index_prev_month == 0:
                self.month_str_var.set(self.month_names[12])
            else:
                self.month_str_var.set(self.month_names[index_current_month - 1])

            self.make_calendar()

        def next_month(self):
            index_current_month = int(self.month_names.index(self.month_str_var.get()))

            #  index 13 does not exist, use index 1 instead, which is January.
            try:
                self.month_str_var.set(self.month_names[index_current_month + 1])
            except IndexError:
                self.month_str_var.set(self.month_names[1])

            self.make_calendar()

        def fill_days(self):
            col = 0
            #  Creates days label
            for day in self.day_names:
                self.lbl_day = tk.Label(self.frame_days, text=day)
                self.lbl_day.grid(row=0, column=col)
                col += 1

        def make_calendar(self):
            #  Delete date buttons if already present.
            #  Each button must have its own instance attribute for this to work.
            try:
                for dates in self.m_cal:
                    for date in dates:
                        if date == 0:
                            continue

                        self.delete_buttons(date)

            except AttributeError:
                pass

            year = int(self.year_str_var.get())
            month = self.month_names.index(self.month_str_var.get())
            self.m_cal = calendar.monthcalendar(year, month)

            #  build date buttons.
            for dates in self.m_cal:
                row = self.m_cal.index(dates) + 1
                for date in dates:
                    col = dates.index(date)

                    if date == 0:
                        continue

                    self.make_button(str(date), str(row), str(col))

        def make_button(self, date, row, column):
            exec(
                "self.btn_" + date + "= ttk.Button(self.frame_days, text=" + date + ", width=5)\n"
                                                                                    "self.btn_" + date + ".grid(row=" + row + " , column=" + column + ")\n"
                                                                                                                                                      "self.btn_" + date + ".bind(\"<Button-1>\", self.get_date)"
            )

        def delete_buttons(self, date):
            exec(
                "self.btn_" + str(date) + ".destroy()"
            )

        def get_date(self, clicked=None):
            clicked_button = clicked.widget
            year = self.year_str_var.get()
            month = self.month_names.index(self.month_str_var.get())
            date = clicked_button['text']
            #  Change string format for different date formats.
            global full_date2
            full_date2 = '%02d-%02d-%s' % (date, month, year)
            fecha2l = ttk.Label(root, text=full_date2).place(x=350, y=370)
            print(full_date2)
            self.destroy()

    if __name__ == '__main__':
        def application2():
            app = MyDatePicker2()

fecha = ttk.Button(root, text="ingrese  fecha desde ", command=application).place(x=200, y=330)
fecha2 = ttk.Button(root, text="ingrese  fecha  hasta", command=application2).place(x=200, y=370)
#########


#excel#\
def llenar():

            desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            test=motivos.get(cm.get())
            print(test)
            shutil.copy2('assets\FORMULARIO DE AUTORIZACIONES.xlsx', 'formulario de autorizaciones.xlsx')
            wb = load_workbook(filename='formulario de autorizaciones.xlsx', data_only=True)

            sheet = wb.active
            img = openpyxl.drawing.image.Image('assets/minilogo.png')
            img.anchor(sheet['B1'])
            sheet.add_image(img)
            now = datetime.datetime.now().strftime("%d-%m-%Y")
            sheet['N8'] = now
            sheet['F11'] = n.get()
            sheet['E10'] = lg.get()
            sheet['E12'] = dp.get()
            sheet['F15'] = test
            sheet['F18'] = full_date
            sheet['R18'] = full_date2
            sheet['F21'] = h1.get()
            sheet['R21'] = h2.get()
            sheet['B24'] = ci.get()

            for i in range(1, 201):
                    if sheet.cell(row=15, column=6).value == sheet.cell(row=i, column=25).value:
                            sheet.cell(row=i, column=25).font= Font(color=colors.RED, bold=True)
                            tkinter.messagebox.showinfo("Atención", "El formulario ha sido guardado en el escritorio ")

            wb.save('formulario de autorizaciones.xlsx')
            shutil.copy('formulario de autorizaciones.xlsx', desktop)
            print("El formulario ha sido guardado en el escritorio ")

            def imprimir():
                os.startfile("formulario de autorizaciones.xlsx", "print")
                tkinter.messagebox.showinfo("Atención", "El formulario esta en cola de impresion ")
            imprimir = ttk.Button(root, text=" Imprimir", command=imprimir).place(x=340, y=410)

            def handler():

                    global file
                    file = open("backup.txt", "r+")
                    x = file.readlines()
                    file.close()
                    if x :
                        if lg.get() == x[0] :
                            root.quit()
                        else:
                                file.close()
                                os.remove("backup.txt")
                                file = open("backup.txt", "w+")
                                file.write(lg.get())
                                file.write("\n")
                                file.write(n.get())
                                file.write("\n")
                                file.write(dp.get())
                                file.close()
                                root.quit()
                    if not x:
                        file.close()
                        h.close()
                        os.remove("backup.txt")
                        file = open("backup.txt", "w+")
                        file.write(lg.get())
                        file.write("\n")
                        file.write(n.get())
                        file.write("\n")
                        file.write(dp.get())
                        file.close()
                        root.quit()
            root.protocol("WM_DELETE_WINDOW", handler)

nombre = ttk.Label(root, text=" Hecho por Alejandro Delgado",width=30).place(x=470, y=430)
llenar = ttk.Button(root, text=" Generar el Formulario", command=llenar).place(x=200, y=410)
root.mainloop()
